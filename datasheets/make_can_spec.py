# -*- coding: utf-8 -*-
"""BMS CAN ID 명세 xlsx 생성.

기존 BMS_CAN_ID_명세_최신_PR31.xlsx 의 시트 구성을 그대로 따르고,
비트맵 시트(내 공부/BMS_CAN_ID_명세.xlsx 에만 있던 것)를 합쳤다.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = r"C:/Users/user/STM32CubeIDE/workspace_1.19.0/BMS_PROJECT/BMS_CAN_ID_명세_최신_PR39.xlsx"

HDR = PatternFill("solid", fgColor="305496")
TITLE_F = Font(bold=True, size=13)
HDR_F = Font(bold=True, color="FFFFFF")
NOTE_F = Font(italic=True, size=9, color="555555")
WARN_F = Font(bold=True, color="C00000")
NEW_F = Font(bold=True, color="006100")

FILL = {
    "PRIMARY": PatternFill("solid", fgColor="DDEBF7"),
    "LEGACY":  PatternFill("solid", fgColor="EDEDED"),
    "DETAIL":  PatternFill("solid", fgColor="FFF2CC"),
    "SD_ONLY": PatternFill("solid", fgColor="E2EFDA"),
    "DIAG":    PatternFill("solid", fgColor="FCE4D6"),
    "RESERVED":PatternFill("solid", fgColor="F2F2F2"),
}
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)


def sheet(wb, name, title, subtitle, header, rows, notes=(), widths=None):
    ws = wb.create_sheet(name)
    ws["A1"] = title
    ws["A1"].font = TITLE_F
    ws["A2"] = subtitle
    ws["A2"].font = NOTE_F
    for c, h in enumerate(header, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.fill, cell.font, cell.border = HDR, HDR_F, THIN
        cell.alignment = Alignment(horizontal="center", vertical="center")
    r = 5
    for row in rows:
        kind = row[0].split("(")[0]
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = THIN
            cell.alignment = Alignment(vertical="center", wrap_text=(c == len(row)))
            if kind in FILL:
                cell.fill = FILL[kind]
        r += 1
    r += 1
    for n in notes:
        ws.cell(row=r, column=1, value=n).font = WARN_F if n.startswith("!!") else NOTE_F
        r += 1
    for i, w in enumerate(widths or [], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A5"
    return ws


wb = openpyxl.Workbook()
wb.remove(wb.active)

SRC = "Source: claude_B.git · Core/Src/B_TEST_BMS_can.c · main @ PR#39"

# ---------------------------------------------------------------- RUN_data
run = [
    ("PRIMARY", "0x040", "BOT", 8, "Global_Fault(u32), BattStat(u16), Alarm(u16)", "팩 전체 고장 코드/배터리 상태/경고 플래그"),
    ("PRIMARY", "0x044", "TOP", 8, "Global_Fault(u32), BattStat(u16), Alarm(u16)", "BOT 0x040과 동일 구조"),
    ("PRIMARY", "0x041", "BOT", 8, "Stack_Voltage(u32,mV), Pack_Voltage(u32,mV)", "셀 스택 전체 전압과 팩 출력단 전압"),
    ("PRIMARY", "0x045", "TOP", 8, "Stack_Voltage(u32,mV), Pack_Voltage(u32,mV)", "BOT 0x041과 동일 구조"),
    ("PRIMARY", "0x042", "BOT", 8, "Pack_Current(i16,10mA), MaxCellVoltage(u16,mV), MinCellVoltage(u16,mV), CellTemp(i16,0.1C)", "팩 충방전 전류(충전=양수)/셀 전압 최대·최소/대표 셀 온도"),
    ("PRIMARY", "0x046", "TOP", 8, "Pack_Current(i16, 항상 0), MaxCellVoltage(u16,mV), MinCellVoltage(u16,mV), CellTemp(i16,0.1C)", "BOT 0x042와 동일 구조. TOP은 SRP/SRN 미사용이라 전류는 항상 0"),
    ("PRIMARY", "0x043", "팩단위", 8, "SOC(u16 permille), FanDuty(u8 %), FanTempEverValid(u8), 예약(4~7)", "★변경★ 팩 SOC + 팬 상태. byte3=0이면 유효 온도 미수신(서미스터 미실장 의심). 보드별 쌍 없음"),
    ("RESERVED", "0x047", "—", "—", "(미사용 — 예약)", "★변경★ 이전에는 0x043과 완전히 동일한 SOC 프레임. 중복이라 삭제하고 자리는 비워 둠. 재사용 금지"),
]
sheet(wb, "RUN_data (일반 CAN)", "BMS CAN ID — RUN data (일반 주기 전송)",
      "BMS_CAN_SendRunData() · 500 kbit/s Classic CAN · " + SRC,
      ["코드 구분", "CAN ID", "보드", "DLC", "필드", "설명"], run,
      notes=["※ RUN data는 0x040~0x046 (BOT 0x040-0x043 / TOP 0x044-0x046). 0x047은 예약(빈 자리).",
             "※ SOC/팬 프레임(0x043)은 팩 단위라 BOT 호출에서 한 번만 송신된다."],
      widths=[12, 10, 8, 6, 62, 70])

# --------------------------------------------------------------- TEST_data
test = [
    ("DETAIL(셀전압)", "0x048-0x04B", "BOT", 8, "Cell[0..15] 전압 (u16x4/frame, mV)", "개별 셀 16개 전압 (4프레임)"),
    ("DETAIL(셀전압)", "0x051-0x054", "TOP", 8, "Cell[0..15] 전압 (u16x4/frame, mV)", "개별 셀 16개 전압 (4프레임)"),
    ("LEGACY", "0x111", "BOT", 8, "Stack_Voltage(u32,mV), Pack_Voltage(u32,mV)", "RUN 전압 프레임과 중복 - 기존 ID 유지"),
    ("LEGACY", "0x211", "TOP", 8, "Stack_Voltage(u32,mV), Pack_Voltage(u32,mV)", "RUN 전압 프레임과 중복 - 기존 ID 유지"),
    ("LEGACY", "0x112", "BOT", 8, "Pack_Current(i16), CC1_Current(i16), CC3_Current(i16), CB_ActiveCells(u16)", "기존 ID 유지"),
    ("LEGACY", "0x212", "TOP", 8, "Pack_Current/CC1/CC3 (전부 항상 0), CB_ActiveCells(u16)", "TOP은 SRP/SRN 미사용이라 전류 필드 전부 0"),
    ("DETAIL(온도)", "0x04C", "BOT", 8, "CellTemp(TS1), FETTemp, IntTemp, CFETOFFTemp (i16x4, 0.1C)", "서미스터별 온도"),
    ("DETAIL(온도)", "0x055", "TOP", 8, "CellTemp(TS1), FETTemp, IntTemp, CFETOFFTemp (i16x4, 0.1C)", "BOT 0x04C와 동일 구조"),
    ("DETAIL(온도)", "0x04D", "BOT", 8, "HDQTemp, MaxCellTemp, MinCellTemp, AvgCellTemp (i16x4, 0.1C)", "HDQ 핀 온도와 셀 온도 최대/최소/평균"),
    ("DETAIL(온도)", "0x056", "TOP", 8, "HDQTemp, MaxCellTemp, MinCellTemp, AvgCellTemp (i16x4, 0.1C)", "BOT 0x04D와 동일 구조"),
    ("LEGACY", "0x120", "BOT", 8, "Global_Fault(u32), BattStat(u16), Alarm(u16)", "RUN 상태 프레임과 중복 - 기존 ID 유지"),
    ("LEGACY", "0x220", "TOP", 8, "Global_Fault(u32), BattStat(u16), Alarm(u16)", "RUN 상태 프레임과 중복 - 기존 ID 유지"),
    ("SD_ONLY(Safety)", "0x05A", "BOT", 8, "SafetyAlert A/B/C, SafetyStatus A/B/C, ProtectionsTriggered, PF_ProtectionsTriggered (u8x8)", "비트 정의는 '비트맵' 시트 참조"),
    ("SD_ONLY(Safety)", "0x06D", "TOP", 8, "SafetyAlert A/B/C, SafetyStatus A/B/C, ProtectionsTriggered, PF_ProtectionsTriggered (u8x8)", "BOT 0x05A와 동일 구조"),
    ("SD_ONLY(PF)", "0x05B", "BOT", 8, "PFAlert A/B/C/D, PFStatus A/B/C/D (u8x8)", "비트 정의는 '비트맵' 시트 참조"),
    ("SD_ONLY(PF)", "0x06E", "TOP", 8, "PFAlert A/B/C/D, PFStatus A/B/C/D (u8x8)", "BOT 0x05B와 동일 구조"),
    ("LEGACY", "0x123", "BOT", 8, "FET_Status, CHG, DSG, PDSG (u8x4), byte4~7=0", "CB_Status1은 0x05C로 분리, AlarmBits는 0x040과 중복이라 삭제"),
    ("LEGACY", "0x223", "TOP", 8, "FET_Status, CHG, DSG, PDSG (u8x4), byte4~7=0", "BOT 0x123과 동일 구조"),
    ("SD_ONLY(밸런싱)", "0x05C", "BOT", 8, "CB_Status1 (u16 비트맵)", "셀 밸런싱 활성 상태 (FET 복합 프레임에서 분리됨)"),
    ("SD_ONLY(밸런싱)", "0x06F", "TOP", 8, "CB_Status1 (u16 비트맵)", "BOT 0x05C와 동일 구조"),
    ("DETAIL(쿨롱)", "0x04E", "BOT", 8, "AccumChargeInt(i32), AccumChargeFrac(u32)", "누적 충전량 정수부/소수부 (userAh, 1 userAh = 10 mAh)"),
    ("DETAIL(쿨롱)", "0x057", "TOP", 8, "AccumChargeInt(i32), AccumChargeFrac(u32)", "BOT 0x04E와 동일 구조 (값은 BOT 것)"),
    ("DETAIL(쿨롱)", "0x04F", "BOT", 8, "AccumChargeTime(u32), CC2Counts(i32)", "누적 적산 시간과 2차 쿨롱 카운터"),
    ("DETAIL(쿨롱)", "0x058", "TOP", 8, "AccumChargeTime(u32), CC2Counts(i32)", "BOT 0x04F와 동일 구조"),
    ("DETAIL(쿨롱)", "0x050", "BOT", 6, "CC3Counts(i32), BattVoltageSum(u16, cV=10mV)", "★변경★ BattVoltageSum 단위 cV 명기(TRM Table 12-23). LD 핀 미사용이라 DLC 6바이트"),
    ("DETAIL(쿨롱)", "0x059", "TOP", 6, "CC3Counts(i32), BattVoltageSum(u16, cV=10mV)", "BOT 0x050과 동일 구조"),
    ("LEGACY", "0x133", "팩단위", 8, "SOC(u16 permille)", "★변경★ TEST용 SOC. RUN 0x043과 같은 이유로 팩 단위라 BOT 호출에서 한 번만 송신"),
    ("RESERVED", "0x233", "—", "—", "(미사용 — 예약)", "★변경★ 이전에는 0x133과 완전히 동일. 중복이라 삭제하고 자리는 비워 둠. 재사용 금지"),
    ("SD_ONLY(CUV)", "0x05D-0x060", "BOT", 8, "CUV 스냅샷[0..15] (u16x4/frame, mV)", "저전압 보호 발동 시점의 셀별 전압 (4프레임)"),
    ("SD_ONLY(CUV)", "0x070-0x073", "TOP", 8, "CUV 스냅샷[0..15] (u16x4/frame, mV)", "BOT과 동일 구조 (4프레임)"),
    ("SD_ONLY(COV)", "0x061-0x064", "BOT", 8, "COV 스냅샷[0..15] (u16x4/frame, mV)", "과전압 보호 발동 시점의 셀별 전압 (4프레임)"),
    ("SD_ONLY(COV)", "0x074-0x077", "TOP", 8, "COV 스냅샷[0..15] (u16x4/frame, mV)", "BOT과 동일 구조 (4프레임)"),
    ("SD_ONLY(밸런싱)", "0x065-0x068", "BOT", 8, "CB_Status2[0..15] (u16x4/frame)", "셀별 밸런싱 상태 2 (4프레임)"),
    ("SD_ONLY(밸런싱)", "0x078-0x07B", "TOP", 8, "CB_Status2[0..15] (u16x4/frame)", "BOT과 동일 구조 (4프레임)"),
    ("SD_ONLY(밸런싱)", "0x069-0x06C", "BOT", 8, "CB_Status3[0..15] (u16x4/frame)", "셀별 밸런싱 상태 3 (4프레임)"),
    ("SD_ONLY(밸런싱)", "0x07C-0x07F", "TOP", 8, "CB_Status3[0..15] (u16x4/frame)", "BOT과 동일 구조 (4프레임)"),
]
sheet(wb, "TEST_data (진단)", "BMS CAN ID — TEST data (진단/테스트 온디맨드)",
      "BMS_CAN_SendTestData() — T_FDCAN_Send_BMS_Data() 호출 시에만. 현재 main 루프에서는 호출되지 않음 · " + SRC,
      ["코드 구분", "CAN ID", "보드", "DLC", "필드", "설명"], test,
      notes=["※ TEST data ID는 0x233(예약)을 제외하면 PR#31 명세와 동일.",
             "※ 색상: 파랑=PRIMARY, 회색=LEGACY, 노랑=DETAIL, 초록=SD_ONLY, 주황=DIAG."],
      widths=[18, 14, 8, 6, 62, 70])

# ---------------------------------------------------------------- DIAG
diag = [
    ("DIAG", "0x300", "팩단위", 8, "TripReason(u8)", "0=정상 1=INIT 2=CHG경로 3=DSG경로 4=통신두절 5=PF 6=스택불일치 7=소프트전압 8=소프트온도 9=재시도초과"),
    ("DIAG", "", "", "", "Flags(u8)", "b0 BOTHOFF어서트 · b1 latch · b2 CHG기대OFF · b3 DSG기대OFF · b4 CHG실측ON · b5 DSG실측ON"),
    ("DIAG", "", "", "", "RetryCount(u8)", "BOTHOFF 해제 재시도 횟수 (한도 3, 초과 시 latch)"),
    ("DIAG", "", "", "", "ChgMismatchCycles(u8)", "CHG 기대=OFF인데 실측=ON인 연속 사이클"),
    ("DIAG", "", "", "", "DsgMismatchCycles(u8)", "DSG 기대=OFF인데 실측=ON인 연속 사이클"),
    ("DIAG", "", "", "", "StackMismatchCycles(u8)", "셀전압 합 vs 스택전압 괴리 연속 사이클"),
    ("DIAG", "", "", "", "TopCommFailCycles(u8)", "TOP 연속 통신 실패 사이클"),
    ("DIAG", "", "", "", "BotCommFailCycles(u8)", "BOT 연속 통신 실패 사이클"),
]
sheet(wb, "DIAG (보호 감시자)", "BMS CAN ID — 진단 프레임 (신규)",
      "BMS_CAN_SendProtectDiag() · 매 루프 송신 · " + SRC,
      ["코드 구분", "CAN ID", "보드", "DLC", "바이트 (0→7 순서)", "설명"], diag,
      notes=["!! 읽는 법: Flags의 b2와 b4가 동시에 1이면 CHG 차단 경로(BOT.DCHG→G3VM→TOP.CFETOFF) 고장. b3·b5는 DSG 경로.",
             "!! 0x300은 '0x310보다 작은 ID' 지시에 맞춰 그 바로 아래로 잡은 값이다. 차량 전체 CAN 할당표와 대조 필요.",
             "※ 팩 단위라 BOT/TOP 쌍이 없다. 이전 위치는 0x080이었다."],
      widths=[12, 10, 8, 6, 26, 100])

# ---------------------------------------------------------------- 변경점
chg = [
    ("변경", "SOC/팬 (RUN, 팩)", "0x043 + 0x047", "0x043", "중복 제거", "두 프레임이 바이트 단위로 완전히 동일했다. SOC는 32S 직렬 스트링 하나의 속성이라 보드별로 나눌 수 없다(모든 셀에 같은 전류). 0x043만 송신."),
    ("변경", "0x047", "SOC (TOP)", "예약(미사용)", "삭제 후 빈 자리", "뒤쪽 ID를 당기면 배포된 명세가 전부 어긋나므로 자리는 비워 둔다. 재사용 금지."),
    ("변경", "0x043 byte 2·3", "0 (미사용)", "FanDuty / FanTempEverValid", "필드 추가", "byte 2~7이 0으로 낭비되던 자리. 팬은 이 보드 최대 소비원(100%=8.52W, MCU는 8.6mW)인데 밖에서 관측 불가였다. byte0~1만 읽던 수신기는 영향 없음."),
    ("변경", "SOC (TEST, 팩)", "0x133 + 0x233", "0x133", "중복 제거", "RUN 0x043/0x047과 완전히 같은 사유. 두 프레임이 바이트 단위로 동일했다. 0x233은 예약으로 비워 둔다."),
    ("변경", "보호 감시자 진단", "0x080", "0x300", "ID 이동", "BMS 측정 데이터 블록(0x040~0x07F)과 성격이 다르고, CAN은 ID가 낮을수록 우선순위가 높아 진단 프레임이 실측 데이터보다 앞설 이유가 없다. '0x310보다 작게' 지시 반영."),
    ("신규", "0x300 자체", "(없음)", "DIAG 프레임", "신설", "MCU 보호 감시자(B_BMS_protect.c) 상태. PR#31 명세에는 이 프레임이 없었다."),
    ("변경", "0x050 / 0x059", "BattVoltageSum (단위 미기재)", "BattVoltageSum (cV=10mV)", "단위 명기", "TRM Table 12-23 확인. 값·구조·DLC(6바이트) 전부 불변. 문서상 명확화만."),
    ("정정", "송신 주기", "약 100 ms", "약 300 ms (추정)", "서술 정정", "HAL_Delay(63)은 사이클 끝에 붙는 값일 뿐. 실제 주기는 I2C가 지배 — DirectCommands()가 호출마다 delayUS(2000), 보드당 약 40회 → 두 보드 ~168 ms + 100kHz 전송 ~68 ms. EVM_TEST.md 6절에서 실측 필요."),
    ("동일", "TEST data 36행", "0x048~0x07F, 0x1xx/0x2xx", "동일", "변경 없음", "PR#31 명세와 ID·보드·DLC 전부 일치."),
    ("동일", "RUN data 6행", "0x040~0x042, 0x044~0x046", "동일", "변경 없음", "STATUS/VOLTAGE/CURRENT는 BOT·TOP 모두 그대로."),
    ("동일", "물리 계층", "500 kbit/s, 빅엔디안, 재전송 DISABLE", "동일", "변경 없음", "비트레이트 계산 검증: PLL1Q 80MHz / (Prescaler 10 x 16 Tq) = 500 kbit/s."),
]
sheet(wb, "변경점(PR31 대비)", "PR#31 명세 대비 변경점",
      "기준: BMS_CAN_ID_명세_최신_PR31.xlsx (B_TEST_BMS_can.c @ b41d2fe) → 현재 main @ PR#39",
      ["구분", "항목", "이전", "현재", "종류", "사유 / 설명"], chg,
      notes=["※ ID 재배치는 없다. RUN 8행 중 1행 삭제(0x047), TEST 36행 전부 동일, 신규 1행(0x300).",
             "※ PR#32~#38 사이에 생긴 변경 전부를 담고 있다."],
      widths=[8, 22, 30, 28, 14, 95])

# ---------------------------------------------------------------- 비트맵
BITS = [
    ("Safety Status A (BQ 0x03) → 0x05A/0x06D byte3 · Alert → byte0 · Global_Fault byte3", "", "", "", ""),
    ("SafetyStatusA", "2", "0x04", "CUV", "✓ Cell Undervoltage — 셀 저전압"),
    ("SafetyStatusA", "3", "0x08", "COV", "✓ Cell Overvoltage — 셀 과전압"),
    ("SafetyStatusA", "4", "0x10", "OCC", "✓ Overcurrent in Charge — 충전 과전류"),
    ("SafetyStatusA", "5", "0x20", "OCD1", "✓ Overcurrent in Discharge 1 — 방전 과전류 1단"),
    ("SafetyStatusA", "6", "0x40", "OCD2", "✓ Overcurrent in Discharge 2 — 방전 과전류 2단"),
    ("SafetyStatusA", "7", "0x80", "SCD", "✓ Short Circuit in Discharge — 방전 단락"),
    ("Safety Status B (BQ 0x05) → 0x05A/0x06D byte4 · Alert → byte1 · Global_Fault byte2", "", "", "", ""),
    ("SafetyStatusB", "0", "0x01", "UTC", "✓ Undertemperature in Charge — 충전 저온"),
    ("SafetyStatusB", "1", "0x02", "UTD", "✓ Undertemperature in Discharge — 방전 저온"),
    ("SafetyStatusB", "2", "0x04", "UTINT", "✓ Internal Undertemperature — 내부 저온"),
    ("SafetyStatusB", "4", "0x10", "OTC", "✓ Overtemperature in Charge — 충전 과온"),
    ("SafetyStatusB", "5", "0x20", "OTD", "✓ Overtemperature in Discharge — 방전 과온"),
    ("SafetyStatusB", "6", "0x40", "OTINT", "✓ Internal Overtemperature — 내부 과온"),
    ("SafetyStatusB", "7", "0x80", "OTF", "✓ Overtemperature FET — FET 과온"),
    ("Safety Status C (BQ 0x07) → 0x05A/0x06D byte5 · Alert → byte2 · Global_Fault byte1", "", "", "", ""),
    ("SafetyStatusC", "1", "0x02", "HWDF", "✓ Host Watchdog Fault — 호스트 워치독 결함"),
    ("SafetyStatusC", "2", "0x04", "PTO", "Precharge Timeout (코드 미파싱)"),
    ("SafetyStatusC", "4", "0x10", "COVL", "✓ Cell Overvoltage Latch — 과전압 래치"),
    ("SafetyStatusC", "5", "0x20", "OCDL", "✓ Overcurrent Discharge Latch — 방전 과전류 래치"),
    ("SafetyStatusC", "6", "0x40", "SCDL", "✓ Short Circuit Discharge Latch — 방전 단락 래치"),
    ("SafetyStatusC", "7", "0x80", "OCD3", "✓ Overcurrent in Discharge 3 — 방전 과전류 3단"),
    ("PF Status A (BQ 0x0B) → 0x05B/0x06E byte4 · PF Alert A → byte0", "", "", "", ""),
    ("PFStatusA", "0", "0x01", "SUV", "✓ Safety Cell Undervoltage PF — 영구 저전압"),
    ("PFStatusA", "1", "0x02", "SOV", "✓ Safety Cell Overvoltage PF — 영구 과전압"),
    ("PFStatusA", "2", "0x04", "SOCC", "Safety Overcurrent Charge PF (코드 미파싱)"),
    ("PFStatusA", "3", "0x08", "SOCD", "Safety Overcurrent Discharge PF (코드 미파싱)"),
    ("PFStatusA", "4", "0x10", "SOT", "Safety Overtemperature PF (코드 미파싱)"),
    ("PFStatusA", "5", "0x20", "SOTF", "Safety Overtemperature FET PF (코드 미파싱)"),
    ("PFStatusA", "7", "0x80", "CUDEP", "Copper Deposition PF (코드 미파싱)"),
    ("PF Status B (BQ 0x0C) → 0x05B/0x06E byte5 · PF Alert B → byte1", "", "", "", ""),
    ("PFStatusB", "0", "0x01", "CFETF", "Charge FET PF — BOT에서 '껐는데 전류가 흐른다'를 잡는다 (임계 0.1A/5s)"),
    ("PFStatusB", "1", "0x02", "DFETF", "Discharge FET PF (임계 -0.5A/5s)"),
    ("PFStatusB", "2", "0x04", "2LVL", "Second Level Protector PF"),
    ("PFStatusB", "3", "0x08", "VIMR", "✓ Voltage Imbalance at Rest — 휴지 중 전압 불균형"),
    ("PFStatusB", "4", "0x10", "VIMA", "✓ Voltage Imbalance Active — 동작 중 전압 불균형"),
    ("PFStatusB", "5", "0x20", "VSSF", "Voltage Sense Self-Test Fail"),
    ("PF Status C/D (BQ 0x0D/0x0E) → 0x05B/0x06E byte6·byte7 — 코드는 원시값만 전송", "", "", "", ""),
    ("PFStatusC", "—", "—", "OTPF/DRMF 등", "OTP 메모리/내부 진단 PF 비트 — TRM 참조"),
    ("PFStatusD", "—", "—", "TOSF 등", "Top-of-Stack 기타 PF 비트 — TRM 참조"),
    ("FET Status (BQ 0x7F) → 0x123/0x223 byte0 — TOP 보드 값", "", "", "", ""),
    ("FET_Status", "0", "0x01", "CHG_FET", "✓ Charge FET 상태 (1=ON) — byte1에 별도 전송"),
    ("FET_Status", "1", "0x02", "PCHG_FET", "Precharge FET 상태"),
    ("FET_Status", "2", "0x04", "DSG_FET", "✓ Discharge FET 상태 (1=ON) — byte2에 별도 전송"),
    ("FET_Status", "3", "0x08", "PDSG_FET", "✓ Predischarge FET 상태 — byte3에 별도 전송"),
    ("FET_Status", "4", "0x10", "DCHG_PIN", "DCHG 핀 상태 (BOT에서는 이 핀이 TOP.CFETOFF를 구동)"),
    ("FET_Status", "5", "0x20", "DDSG_PIN", "DDSG 핀 상태 (BOT에서는 이 핀이 TOP.DFETOFF를 구동)"),
    ("FET_Status", "6", "0x40", "ALRT_PIN", "ALERT 핀 상태"),
    ("Alarm Status (BQ 0x62, 16bit) / Alarm Raw (0x64) → 0x040/0x044 byte6-7", "", "", "", ""),
    ("AlarmBits(16)", "7", "0x0080", "FULLSCAN", "✓ ADC 풀스캔 완료 — 코드가 이 비트로 전압/전류 갱신을 게이트한다"),
    ("AlarmBits(16)", "13", "0x2000", "SSBC", "Safety Status B/C 계열 알람 활성"),
    ("AlarmBits(16)", "14", "0x4000", "SSA", "Safety Status A 계열 알람 활성"),
    ("AlarmBits(16)", "15", "0x8000", "SSBC_HI", "상위 Safety/PF 알람 (칩 설정에 따라 다름)"),
    ("Battery Status (BQ 0x12, 16bit) → 0x040/0x044 byte4-5 — 코드는 원시값만 전송", "", "", "", ""),
    ("BattStat(16)", "0", "0x0001", "CFGUPDATE", "Config Update 모드 진입 중"),
    ("BattStat(16)", "2", "0x0004", "PCHG_MODE", "Precharge 모드"),
    ("BattStat(16)", "3", "0x0008", "SLEEP_EN", "Sleep 허용됨"),
    ("BattStat(16)", "5", "0x0020", "POR", "Power-On Reset 발생"),
    ("BattStat(16)", "6", "0x0040", "WD", "Watchdog Reset 발생"),
    ("BattStat(16)", "13", "0x2000", "SS", "Safety Status 활성 (보호 동작 중)"),
    ("BattStat(16)", "14", "0x4000", "PF", "Permanent Fail 활성"),
    ("BattStat(16)", "15", "0x8000", "SLEEP", "✓ SLEEP 모드 동작 중 — Enter_Sleep_Sequence()가 이 비트를 본다"),
]
ws = wb.create_sheet("비트맵(Fault_Bits)")
ws["A1"] = "BMS Fault / Status 비트맵"
ws["A1"].font = TITLE_F
ws["A2"] = "출처: 내 공부/BMS_CAN_ID_명세.xlsx 의 비트맵 시트를 옮겨온 것 (해당 파일은 CAN ID 체계가 옛것이라 이 시트만 살림). 기준: BQ76972 TRM + 코드 파싱."
ws["A2"].font = NOTE_F
ws["A3"] = "'코드' 열의 ✓ = B_BMS.c의 BQ769x2_ReadSafetyStatus()/ReadPFStatus()/ReadFETStatus()가 실제로 개별 비트로 파싱하는 것. 나머지는 원시값으로만 전송된다."
ws["A3"].font = NOTE_F
for c, h in enumerate(["레지스터 / 필드", "비트", "마스크", "약어", "의미"], 1):
    cell = ws.cell(row=5, column=c, value=h)
    cell.fill, cell.font, cell.border = HDR, HDR_F, THIN
r = 6
for row in BITS:
    is_hdr = row[1] == "" and row[2] == ""
    for c, v in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.border = THIN
        cell.alignment = Alignment(vertical="center", wrap_text=(c == 5))
        if is_hdr:
            cell.fill = PatternFill("solid", fgColor="D9E1F2")
            cell.font = Font(bold=True)
        elif c == 5 and str(v).startswith("✓"):
            cell.font = NEW_F
    r += 1
r += 1
ws.cell(row=r, column=1, value="※ Global_Fault_Flags (u32) = byte0:0x00 | byte1:SafetyStatusC | byte2:SafetyStatusB | byte3:SafetyStatusA (빅엔디안)").font = NOTE_F
ws.cell(row=r + 1, column=1, value="※ Alarm 비트 의미는 Default Alarm Mask 설정에 따라 매핑이 달라진다. 코드가 확정적으로 쓰는 것은 bit7(FULLSCAN, 0x0080) 뿐이다.").font = NOTE_F
for i, w in enumerate([46, 8, 10, 16, 78], 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
ws.freeze_panes = "A6"

# ---------------------------------------------------------------- 개요
ws = wb.create_sheet("개요", 0)
ws["A1"] = "BMS CAN 통신 명세 (RUN + TEST + DIAG)"
ws["A1"].font = Font(bold=True, size=14)
rows = [
    ("", ""),
    ("■ 물리 계층 / 프레임 형식", ""),
    ("CAN 컨트롤러", "FDCAN1 (PA11=RX, PA12=TX), Classic CAN 모드 (FD 아님)"),
    ("프레임 형식", "표준 11-bit ID · 대부분 DLC=8 · 예외: 0x050 / 0x059 는 DLC=6"),
    ("바이트 순서", "빅엔디안(MSB first) — I2C 원본은 리틀엔디안이지만 CAN에서 변환된다"),
    ("자동 재전송", "DISABLE (AutoRetransmission off)"),
    ("", ""),
    ("■ 비트레이트", ""),
    ("FDCAN 커널 클럭", "PLL1Q = 80 MHz (MSI 48MHz /PLLM3 → 16MHz, ×PLLN10 → 160MHz VCO, /PLLQ2 → 80MHz)"),
    ("비트타이밍", "NominalPrescaler=10, SJW=1, TimeSeg1=12, TimeSeg2=3 → 16 Tq"),
    ("비트레이트", "80 MHz / (10 × 16) = 500 kbit/s (샘플포인트 81.25%)"),
    ("", ""),
    ("■ 송신 주기", ""),
    ("RUN data", "메인 루프마다 — 실측 필요. 추정 약 300 ms (아래 참조)"),
    ("!! 주기 정정", "PR#31 명세는 '약 100 ms'로 적었으나 틀리다. HAL_Delay(63)은 사이클 끝에 붙는 값일 뿐이고, 실제 주기는 I2C가 지배한다: DirectCommands()가 호출마다 delayUS(2000)을 걸고 보드당 ~40회 → 두 보드 ~168 ms, 여기에 100 kHz 실제 전송 ~68 ms. EVM_TEST.md 6절에서 실측할 것."),
    ("TEST data", "T_FDCAN_Send_BMS_Data() 호출 시에만. 현재 main 루프에서는 호출되지 않는다."),
    ("DIAG data", "매 루프 (BMS_CAN_SendProtectDiag())"),
    ("", ""),
    ("■ 공유 하드웨어 주의 (중요)", ""),
    ("전류 데이터 출처", "Pack_Current / CC1 / CC3 / 쿨롱카운터는 board_type과 무관하게 항상 BOT 값. TOP 프레임에 실려도 내용은 BOT의 전류다. TOP은 SRP/SRN 미사용."),
    ("FET 데이터 출처", "FET_Status / CHG / DSG / PDSG / FET_Temp 는 항상 TOP 값."),
    ("SOC / 팬", "팩 단위. 0x043 하나만 송신한다 (이전에는 0x047로도 중복 송신했다)."),
    ("", ""),
    ("■ 단위 / 스케일", ""),
    ("셀 전압 / Max·Min", "uint16, mV (1 LSB = 1 mV)"),
    ("Stack / Pack 전압", "uint32, mV (BQ raw × 10 — DAConfiguration=0x06, USER_VOLTS_CV=1 센티볼트)"),
    ("BattVoltageSum", "uint16, cV (10 mV) — TRM Table 12-23"),
    ("전류 (Pack/CC1/CC3)", "int16, 10 mA (USER_AMPS=2 센티암페어). 충전=양수, 방전=음수. |1000| = 10 A"),
    ("누적 충전량", "userAh — 1 userAh = 10 mAh (USER_AMPS 설정에 연동)"),
    ("온도 전체", "int16, 0.1 °C (값 = ℃ × 10), 부호 있음"),
    ("SOC", "uint16 permille (1000 = 100.0 %)"),
    ("팬 duty", "uint8, % (0~100)"),
    ("", ""),
    ("■ ID 대역", ""),
    ("0x040 ~ 0x046", "RUN data (BOT 0x040-0x043 / TOP 0x044-0x046)"),
    ("0x047", "예약 — 비어 있음. 재사용 금지 (이전 TOP SOC, 중복이라 삭제)"),
    ("0x048 ~ 0x07F", "TEST data (DETAIL / SD_ONLY)"),
    ("0x080", "비어 있음 — 이전 DIAG 위치. 0x300으로 이동함"),
    ("0x233", "예약 — 비어 있음. 재사용 금지 (이전 TEST TOP SOC, 중복이라 삭제)"),
    ("0x111 / 0x112 / 0x120 / 0x123 / 0x133 (+ TOP 0x2xx)", "LEGACY TEST 프레임 — 기존 ID 유지"),
    ("0x300", "DIAG — MCU 보호 감시자 상태 (신규)"),
    ("", ""),
    ("!! 확인 필요", "0x300은 '0x310보다 작은 ID' 지시에 맞춰 잡은 값이다. 차량 전체 CAN 할당표와 대조해 다른 노드와 겹치지 않는지 확인할 것. CAN은 ID가 낮을수록 우선순위가 높다."),
    ("", ""),
    ("Source", "claude_B.git · Core/Src/B_TEST_BMS_can.c · main @ PR#39"),
]
r = 3
for k, v in rows:
    ws.cell(row=r, column=1, value=k)
    ws.cell(row=r, column=2, value=v).alignment = Alignment(wrap_text=True, vertical="center")
    if k.startswith("■"):
        ws.cell(row=r, column=1).font = Font(bold=True, size=11)
    elif k.startswith("!!"):
        ws.cell(row=r, column=1).font = WARN_F
        ws.cell(row=r, column=2).font = WARN_F
    r += 1
ws.column_dimensions["A"].width = 46
ws.column_dimensions["B"].width = 118

wb.save(OUT)
print("saved:", OUT)
print("sheets:", wb.sheetnames)
